import openpyxl
import csv

# パスの定義
static_path = "static"
search_method_excel_path = static_path + "/input_excel/search_method.xlsx"
output_reins_excel_path = static_path + "/output_excel/output_reins.xlsx"
output_reins_csv_path = static_path + "/csv/output_reins.csv"


def excel_to_list(input_excel_path: str = "input.xlsx"):
    workbook = openpyxl.load_workbook(input_excel_path)
    sheet = workbook.active
    row_num = sheet.max_row
    col_num = sheet.max_column
    data_list = []
    for row in range(1, row_num+1):
        row_data = []
        for col in range(1, col_num+1):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(cell_value)
        data_list.append(row_data)
    return data_list

def list_to_excel(to_excel_list: list , output_excel_path: str = "output.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    row_num = len(to_excel_list)
    col_num = len(to_excel_list[0])
    for row in range(row_num):
        for col in range(col_num):
            sheet.cell(row=row+1, column=col+1).value = to_excel_list[row][col]
    workbook.save(output_excel_path)
    
def csv_to_list(csv_path: str = "output.csv"):
    """ 多次元データを含むcsvからリストに変換 """
    data_list = []
    with open(csv_path, 'r' , encoding="utf-8-sig") as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            data_list.append(row)
    return data_list


def csv_to_excel(input_csv_path, output_excel_path):
    """ csvファイルをExcelファイルに変換する関数 """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # CSVファイルを開き、行ごとにExcelシートに書き込む
    with open(input_csv_path, 'r', newline='', encoding='utf-8') as csvfile:
        csv_reader = csv.reader(csvfile)
        for row_index, row in enumerate(csv_reader, start=1):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)
    # Excelファイルに保存
    workbook.save(output_excel_path)


def main():
    csv_to_excel(output_reins_csv_path , output_reins_excel_path)
    excel_list = excel_to_list(output_reins_excel_path)
    list_to_excel(excel_list , output_reins_excel_path)


if __name__ == "__main__":
    main()