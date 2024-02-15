from fastapi import FastAPI
from pydantic import BaseModel
import numpy as np
import re
import time
import os
import csv
import openpyxl

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication


# パスの定義
static_path = "static"
mail_excel_path = static_path + "/input_excel/email_pw.xlsx"
search_method_excel_path = static_path + "/input_excel/search_method.xlsx"
output_reins_excel_path = static_path + "/output_excel/output_reins.xlsx"
log_txt_path = static_path + "/log/log.txt"

# 環境変数の取得
user_id , password = os.environ.get('SECRET_USER_ID') , os.environ.get('SECRET_PASSWORD')

class logText:
    def __init__(self , log_txt_path) -> None:
        self.log_txt_path = log_txt_path
        # logの保存ファイルを空にする
        with open(self.log_txt_path, 'w') as file:
            file.write('')

    def add_log_txt(self , add_log_text):
        """ logを付け加える関数 """
        with open(self.log_txt_path, 'a') as file:
            file.write("\n" + add_log_text)
log_txt = logText(log_txt_path)


def send_py_gmail(
    message_subject , message_body , from_email_smtp_password ,
    from_email , to_email , cc_mail_row_list = [] , file_path = "",
):
    """ メールを送信する関数 """
    msg = MIMEMultipart()
    msg['To'] = to_email
    msg['From'] = from_email
    if cc_mail_row_list !=[]:
        msg['Cc'] = ",".join(cc_mail_row_list)
    msg['Subject'] = message_subject
    msg.attach(MIMEText(message_body))
    # ファイルをメールに添付
    file_name = os.path.basename(file_path)
    with open(file_path , "rb") as f:
        attachment = MIMEApplication(f.read())
    attachment.add_header("Content-Disposition", "attachment", filename = file_name)
    msg.attach(attachment)
    # サーバーを指定しメールを送信
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(from_email, from_email_smtp_password)
    server.send_message(msg)
    server.quit()


def mail_list_from_excel(mail_excel_path):
    """ Excelファイルからメールのリストを取得する関数 """
    mail_list = []
    workbook = openpyxl.load_workbook(mail_excel_path)
    log_txt.add_log_txt("Excelのワークブック起動完了 : workbook = openpyxl.load_workbook()")
    sheet = workbook.active
    log_txt.add_log_txt("ワークブックのアクティブ化完了 : sheet = workbook.active")
    receive_email_number = 100
    for index in range(receive_email_number):
        mail = sheet.cell(row = index + 2 , column = 3).value
        # emailかどうかを判定（「@」「.」の有無）
        if mail is not None:
            if '@' in mail and '.' in mail:
                mail_list.append(mail)
        else:
            break
    log_txt.add_log_txt("セルの編集可能が証明 : mail = sheet.cell(row = index + 2 , column = 3).value")
    # ccのメールのリストを取得
    cc_mail_list = []
    for index in range(len(mail_list)):
        # D列以降を判定
        cc_mail_row_list = []
        for col in range(receive_email_number):
            cc_mail = sheet.cell(row = index + 2 , column = 4 + col).value
            # emailかどうかを判定（「@」「.」の有無）
            if cc_mail is not None:
                if '@' in cc_mail and '.' in cc_mail:
                    cc_mail_row_list.append(cc_mail)
            else:
                break
        cc_mail_list.append(cc_mail_row_list)
    # 送信元メールアドレスとアプリパスワードを取得
    from_email = sheet.cell(row = 2 , column = 1).value
    from_email_smtp_password = sheet.cell(row = 2 , column = 2).value
    return mail_list , cc_mail_list , from_email , from_email_smtp_password


def excel_to_list(input_excel_path: str = "input.xlsx"):
    workbook = openpyxl.load_workbook(input_excel_path)
    log_txt.add_log_txt("Excelのワークブック起動完了 : workbook = openpyxl.load_workbook()")
    sheet = workbook.active
    log_txt.add_log_txt("ワークブックのアクティブ化完了 : sheet = workbook.active")
    row_num = sheet.max_row
    log_txt.add_log_txt(f"row_num : {row_num}")
    col_num = sheet.max_column
    log_txt.add_log_txt(f"col_num : {col_num}")
    data_list = []
    for row in range(1, row_num+1):
        row_data = []
        for col in range(1, col_num+1):
            cell_value = sheet.cell(row=row, column=col).value
            log_txt.add_log_txt(f"cell_value : {cell_value}")
            log_txt.add_log_txt(f"row , col : {row} , {col} \n")
            row_data.append(cell_value)
        data_list.append(row_data)
    log_txt.add_log_txt("セルの編集可能が証明 : cell_value = sheet.cell(row=row, column=col).value")
    return data_list


def list_to_excel(to_excel_list: list , output_excel_path: str = "output.xlsx"):
    workbook = openpyxl.Workbook()
    log_txt.add_log_txt("Excelのワークブック起動完了 : workbook = openpyxl.load_workbook()")
    sheet = workbook.active
    log_txt.add_log_txt("ワークブックのアクティブ化完了 : sheet = workbook.active")
    row_num = len(to_excel_list)
    col_num = len(to_excel_list[0])
    for row in range(row_num):
        for col in range(col_num):
            sheet.cell(row=row+1, column=col+1).value = to_excel_list[row][col]
    log_txt.add_log_txt("セルの編集可能が証明 : sheet.cell(row=row+1, column=col+1).value = to_excel_list[row][col]")
    workbook.save(output_excel_path)


def remove_non_number(text):
    divided_number = re.findall(r'\d+', text)  # 文字列から数字にマッチするものをリストとして取得
    integrate_only_number = ''.join(divided_number)
    integrate_only_number = re.sub(r'\D', '', text)  # 元の文字列から数字以外を削除＝数字を抽出
    return divided_number , integrate_only_number

def csv_to_excel(input_csv_path, output_excel_path):
    """ csvファイルをExcelファイルに変換する関数 """
    workbook = openpyxl.Workbook()
    log_txt.add_log_txt("Excelのワークブック起動完了 : workbook = openpyxl.load_workbook()")
    sheet = workbook.active
    log_txt.add_log_txt("ワークブックのアクティブ化完了 : sheet = workbook.active")
    # CSVファイルを開き、行ごとにExcelシートに書き込む
    with open(input_csv_path, 'r', newline='', encoding='utf-8') as csvfile:
        csv_reader = csv.reader(csvfile)
        log_txt.add_log_txt("csvファイルの起動成功 : csv_reader = csv.reader(csvfile)")
        for row_index, row in enumerate(csv_reader, start=1):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)
    log_txt.add_log_txt("セルの編集可能が証明 : sheet.cell(row=row_index, column=col_index, value=value)")
    # Excelファイルに保存
    workbook.save(output_excel_path)
    log_txt.add_log_txt("Excelファイルに保存が成功")

def list_to_csv(to_csv_list: list , csv_path: str = "output.csv"):
    """ 多次元リストのデータをcsvファイルに保存する関数 """
    with open(csv_path, 'w' , encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(to_csv_list)


def csv_to_list(csv_path: str = "output.csv"):
    """ 多次元データを含むcsvからリストに変換 """
    data_list = []
    with open(csv_path, 'r' , encoding="utf-8-sig") as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            data_list.append(row)
    return data_list



def get_search_option(input_csv_path):
    """ 定期実行ツールがcsvファイルから検索方法と条件を取得する関数 """
    search_option_list = csv_to_list(input_csv_path)
    search_method_value = search_option_list[1][0]
    search_requirement = int( search_option_list[1][1] )
    return search_method_value , search_requirement




class RequestData(BaseModel):
    data_list: list
    data_str: str
    data_int: int

app = FastAPI()

@app.post("/")
def cloud_fast_api_1(data: RequestData):
    # 呼び出し元からデータを取得
    data_list = data.data_list
    data_str = data.data_str
    data_int = data.data_int

    log_txt.add_log_txt("API処理開始")

    # メールアドレスのリストをExcelから取得
    mail_list , cc_mail_list , from_email , from_email_smtp_password = mail_list_from_excel(mail_excel_path)
    log_txt.add_log_txt("メールアドレスのリストをExcelから取得 が成功")
    
    try:
        # スクレイピング結果のリストをExcelファイルに保存
        output_excel_list = excel_to_list(output_reins_excel_path)
        list_to_excel(output_excel_list , output_reins_excel_path)
        ##### 最終的にはExcelの定型フォームに貼り付け
        
        log_txt.add_log_txt("スクレイピング結果のcsvファイルをExcelファイルに変更 : 完了")
        # メールの送信文
        message_subject = "RenderのExcel操作のAPIテスト"
        message_body = f"""
            RenderでExcel操作の達成完了しました。
        """
        # file_path = mail_excel_path
        file_path = output_reins_excel_path
    
    except:
        # メールの送信文
        message_subject = "RenderのExcel操作のAPIテスト"
        message_body = f"""
            RenderでのExcelを操作に不備がありました。
        """
        file_path = log_txt_path

    # 全てのメールにスクレイピング結果のExcelを送信
    for loop , to_email in enumerate(mail_list):
        cc_mail_row_list = cc_mail_list[loop]
        send_py_gmail(
            message_subject , message_body , from_email_smtp_password ,
            from_email , to_email , cc_mail_row_list = cc_mail_row_list ,
            file_path = file_path ,
        )


    return {"api_output_list": output_excel_list}